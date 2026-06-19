import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_tracker():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CT-Buddy LLM Trials'
    headers = ['Date','Project Name','LLM','Model Version','Low CT Dimensions',
               'Prompt Summary','Response (first 200 chars)','Stayed Socratic? Y/N',
               'Leaked Answer? Y/N','Quality 1-5','Input Tokens (est)','Output Tokens (est)',
               'Cost Est ($)','Notes']
    fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    font = Font(bold=True, color='FFFFFF', name='Arial')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal='center', wrap_text=True)
    widths = [12,20,10,16,22,28,38,16,16,10,14,14,12,28]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.append([
        datetime.now().strftime('%Y-%m-%d'),
        'Platformer game (My First Project)',
        'Claude','claude-sonnet-4',
        'Logic: 0/4, Math Operators: 0/4',
        'Socratic tutor, 0 on Logic, ask one guiding question',
        'Hey there! When your player touches a hazard...',
        'Y','N',4,'~350','~120','~$0.00062',
        'Good. Warm tone. Project-specific. One question. Possibly too long.'
    ])
    wb.save('CT_Buddy_Trial_Tracker.xlsx')
    print('Tracker created: CT_Buddy_Trial_Tracker.xlsx')

create_tracker()
